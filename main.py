from __future__ import print_function
import pickle
import os.path
from googleapiclient import errors
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import re
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import base64

SCOPES = ['https://mail.google.com/']

def main():
    creds = None
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('gmail', 'v1', credentials=creds)
    com1 = Comunicador(service)
    com1.iterarPorDiccionarios()



class Comunicador:
    def __init__(self, service):
        self.service = service
        self.nombreAMostrar = input("INGRESE NOMBRE A MOSTRAR: ")
        self.contactos_path = input("INGRESE RUTA DE EXCEL DE CONTACTOS: ")
        self.contactos = load_workbook(self.contactos_path).active
        self.copia = input('INGRESE EL CORREO AL CUAL DESEA COPIAR: ')
        self.asunto = input('INGRESE ASUNTO DEL CORREO: ')
        self.html_path = input('INGRESE RUTA DEL HTML: ')
        self.accion = int(input(
            "INGRESE 1 PARA PRUEBA/ INGRESE 2 PARA ENVIAR/ INGRESE 3 PARA CANCELAR: "))

    def diccionarioFila(self, n):
        dic = {}
        for col in self.contactos.columns:
            if col[0].value != None:
                dic[str(col[0].value)] = col[n].value
            else:
                break
        return dic

    def iterarPorDiccionarios(self):
        if self.accion == 1:
            dic = {}
            for col in self.contactos.columns:
                if col[0].value != None:
                    dic[str(col[0].value)] = col[1].value
                else:
                    break
            editorHTML = EditorHTML(self.html_path, dic)
            htmlEditado = editorHTML.actualizarHTML()
            testEmail = input("INGRESE EL CORREO AL CUAL QUIERE ENVIAR LA PRUEBA: ")
            mensaje = Mensaje(self.service)
            aux = None
            if dic['ATTACHMENT'] == None:
                pass
            else:
                aux = str(dic['ATTACHMENT']).split(';', 50)
            mensaje.create_message(self.nombreAMostrar, testEmail, self.copia, self.asunto, htmlEditado, aux)
            mensaje.send_message()

        elif self.accion == 2:
            for row in range(1, self.contactos.max_row - 1):
                dic = {}
                for col in self.contactos.columns:
                    if col[0].value != None:
                        dic[str(col[0].value)] = col[row].value
                    else:
                        break
                editorHTML = EditorHTML(self.html_path, dic)
                htmlEditado = editorHTML.actualizarHTML()
                mensaje = Mensaje(self.service)
                print(dic)
                aux = None
                if dic['ATTACHMENT'] == None:
                    pass
                else:
                    aux = str(dic['ATTACHMENT']).split(';', 50)

                mensaje.create_message(self.nombreAMostrar, dic['EMAIL'], self.copia, self.asunto, htmlEditado, aux)
                mensaje.send_message()
        else:
            print("CANCELADO SATISFACTORIAMENTE")
            pass


class EditorHTML:
    def __init__(self, html_path, valores):
        self.html = BeautifulSoup(open(html_path, 'r').read(), features="html.parser")
        self.valores = valores

    def actualizarHTML(self):
        for key in self.valores.keys():
            target = self.html.find_all(text=re.compile(r'{}'.format(key)))
            for v in target:
                if self.valores[key] == None:
                    v.replace_with(v.replace(key, ""))
                else:
                    v.replace_with(v.replace(key, self.valores[key]))
        return self.html

class Mensaje():
    def __init__(self, service):
        self.service = service
        self.message = None

    def send_message(self):
        try:
            message = (self.service.users().messages().send(userId="me", body=self.message)
                       .execute())
            print('Message Id: %s' % message['id'])
            return message
        except errors.HttpError as error:
            print('An error occurred: %s' % error)


    def create_message(self, sender, to, copia ,subject, html_Editado, atachmentArray):
        
        message = MIMEMultipart()
        message['To'] = to
        message['From'] = sender
        message['Subject'] = subject
        message['CC'] = copia
        mensajeHTML = MIMEText(html_Editado,'html')
        message.attach(mensajeHTML)

        if atachmentArray != None:
            for file in atachmentArray:
                with open(file, "rb") as f:
                    attach = MIMEApplication(f.read(), _subtype="pdf")
                attach.add_header('Content-Disposition', 'attachment', filename=str(file))
                message.attach(attach)

        self.message = {'raw': base64.urlsafe_b64encode(message.as_bytes()).decode()}

if __name__ == '__main__':
    main()